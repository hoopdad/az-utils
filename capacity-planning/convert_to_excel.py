#!/usr/bin/env python3
"""Convert capacity planning JSON reports into a multi-tab Excel workbook."""

import json
import os
import sys
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# Style constants
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
CRIT_FILL = PatternFill(start_color="F4CCCC", end_color="F4CCCC", fill_type="solid")
GOOD_FILL = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="BFBFBF"),
    right=Side(style="thin", color="BFBFBF"),
    top=Side(style="thin", color="BFBFBF"),
    bottom=Side(style="thin", color="BFBFBF"),
)


def style_header_row(ws, row, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
        cell.border = THIN_BORDER


def style_data_cell(ws, row, col, border=True):
    cell = ws.cell(row=row, column=col)
    if border:
        cell.border = THIN_BORDER
    return cell


def auto_width(ws, min_width=10, max_width=50):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_width), max_width)


def load_json(path):
    try:
        with open(path, "r") as f:
            return json.load(f)
    except Exception:
        return None


def find_sub_dirs(report_dir):
    """Find subscription directories in a report. Returns list of (name, path) tuples."""
    subs = []
    report = Path(report_dir)

    # Check if this is a multi-sub report (subdirectories with JSON files)
    # or single-sub (JSON files directly in report_dir)
    direct_json = list(report.glob("service-inventory.json"))
    if direct_json:
        # Single sub - use directory name as sub name
        return [(report.name, str(report))]

    # Multi-sub - each subdirectory is a subscription
    for d in sorted(report.iterdir()):
        if d.is_dir() and (d / "service-inventory.json").exists():
            subs.append((d.name, str(d)))

    return subs


def build_overview_sheet(wb, subs_data):
    """Tab 1: Overview - one row per subscription with key metrics."""
    ws = wb.active
    ws.title = "Overview"

    headers = [
        "Subscription",
        "Total Resources",
        "Resource Types",
        "Regions Used",
        "Resource Groups",
        "Quotas Checked",
        "Quotas >80%",
        "Quotas >90%",
        "Resources Monitored",
        "High Util (P95≥80%)",
        "Reservations",
        "RI Expiring <90d",
    ]

    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "B2"

    row = 2
    for sub_name, data in subs_data:
        ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER

        inv = data.get("service-inventory")
        if inv and inv.get("summary"):
            s = inv["summary"]
            ws.cell(row=row, column=2, value=s.get("totalResources", 0)).border = THIN_BORDER
            ws.cell(row=row, column=3, value=s.get("uniqueResourceTypes", 0)).border = THIN_BORDER
            ws.cell(row=row, column=4, value=s.get("uniqueRegions", 0)).border = THIN_BORDER
            ws.cell(row=row, column=5, value=s.get("uniqueResourceGroups", 0)).border = THIN_BORDER

        quota = data.get("quota-usage")
        if quota and quota.get("summary"):
            s = quota["summary"]
            ws.cell(row=row, column=6, value=s.get("totalQuotasChecked", 0)).border = THIN_BORDER
            c80 = ws.cell(row=row, column=7, value=s.get("quotasAbove80Percent", 0))
            c80.border = THIN_BORDER
            if c80.value and c80.value > 0:
                c80.fill = WARN_FILL
            c90 = ws.cell(row=row, column=8, value=s.get("quotasAbove90Percent", 0))
            c90.border = THIN_BORDER
            if c90.value and c90.value > 0:
                c90.fill = CRIT_FILL

        trends = data.get("usage-trends")
        if trends and trends.get("summary"):
            s = trends["summary"]
            ws.cell(row=row, column=9, value=s.get("totalResourcesAnalyzed", 0)).border = THIN_BORDER
            recs = trends.get("records", [])
            high_count = sum(1 for r in recs if r.get("unit") == "Percent" and (r.get("p95") or 0) >= 80)
            c_hu = ws.cell(row=row, column=10, value=high_count)
            c_hu.border = THIN_BORDER
            if high_count > 0:
                c_hu.fill = WARN_FILL

        ri = data.get("reserved-instances")
        if ri and ri.get("summary"):
            s = ri["summary"]
            ws.cell(row=row, column=11, value=s.get("totalReservations", 0)).border = THIN_BORDER
            exp = s.get("expiringWithin90Days", 0)
            c_exp = ws.cell(row=row, column=12, value=exp)
            c_exp.border = THIN_BORDER
            if exp and exp > 0:
                c_exp.fill = CRIT_FILL

        row += 1

    auto_width(ws)


def build_inventory_sheet(wb, subs_data):
    """Tab 2: All resources across all subscriptions."""
    ws = wb.create_sheet("Resource Inventory")

    headers = ["Subscription", "Resource Name", "Resource Type", "Region", "Resource Group", "SKU"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    for sub_name, data in subs_data:
        inv = data.get("service-inventory")
        if not inv or not inv.get("records"):
            continue
        for rec in inv["records"]:
            ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=rec.get("name", "")).border = THIN_BORDER
            ws.cell(row=row, column=3, value=rec.get("type", "")).border = THIN_BORDER
            ws.cell(row=row, column=4, value=rec.get("location", "")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=rec.get("resourceGroup", "")).border = THIN_BORDER
            sku = rec.get("sku")
            if isinstance(sku, dict):
                sku = sku.get("name", str(sku))
            elif sku is None:
                sku = ""
            ws.cell(row=row, column=6, value=str(sku)).border = THIN_BORDER
            row += 1

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


def build_inventory_summary_sheet(wb, subs_data):
    """Tab 3: Resource counts by subscription and type (pivot-style)."""
    ws = wb.create_sheet("Resource Summary")

    # Aggregate: {sub -> {type -> count}}
    agg = {}
    all_types = set()
    for sub_name, data in subs_data:
        inv = data.get("service-inventory")
        if not inv:
            continue
        top_types = inv.get("summary", {}).get("topResourceTypes", [])
        if top_types:
            agg[sub_name] = {t["type"]: t["count"] for t in top_types}
            all_types.update(t["type"] for t in top_types)

    if not all_types:
        ws.cell(row=1, column=1, value="No inventory data available")
        return

    sorted_types = sorted(all_types)

    # Headers
    ws.cell(row=1, column=1, value="Subscription")
    for col, t in enumerate(sorted_types, 2):
        ws.cell(row=1, column=col, value=t)
    ws.cell(row=1, column=len(sorted_types) + 2, value="Total")
    style_header_row(ws, 1, len(sorted_types) + 2)
    ws.freeze_panes = "B2"

    row = 2
    for sub_name in sorted(agg.keys()):
        ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
        total = 0
        for col, t in enumerate(sorted_types, 2):
            val = agg[sub_name].get(t, 0)
            total += val
            c = ws.cell(row=row, column=col, value=val if val else "")
            c.border = THIN_BORDER
        ws.cell(row=row, column=len(sorted_types) + 2, value=total).border = THIN_BORDER
        row += 1

    auto_width(ws, min_width=6, max_width=40)


def build_quota_sheet(wb, subs_data):
    """Tab 4: All quota records across subscriptions."""
    ws = wb.create_sheet("Quota Usage")

    headers = ["Subscription", "Provider", "Region", "Quota Name", "Limit", "Current Usage", "% Used", "Status"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    for sub_name, data in subs_data:
        quota = data.get("quota-usage")
        if not quota or not quota.get("records"):
            continue
        for rec in quota["records"]:
            usage_pct = rec.get("usagePercent", 0) or 0
            # Skip zero-usage quotas to keep the sheet manageable
            if usage_pct == 0 and rec.get("currentUsage", 0) == 0:
                continue

            ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=rec.get("provider", "")).border = THIN_BORDER
            ws.cell(row=row, column=3, value=rec.get("region", "")).border = THIN_BORDER
            ws.cell(row=row, column=4, value=rec.get("quotaName", "")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=rec.get("limit", 0)).border = THIN_BORDER
            ws.cell(row=row, column=6, value=rec.get("currentUsage", 0)).border = THIN_BORDER

            pct_cell = ws.cell(row=row, column=7, value=round(usage_pct, 1))
            pct_cell.border = THIN_BORDER
            pct_cell.number_format = "0.0"

            if usage_pct >= 90:
                status = "CRITICAL"
                pct_cell.fill = CRIT_FILL
                ws.cell(row=row, column=8, value=status).fill = CRIT_FILL
            elif usage_pct >= 80:
                status = "WARNING"
                pct_cell.fill = WARN_FILL
                ws.cell(row=row, column=8, value=status).fill = WARN_FILL
            else:
                status = "OK"
            ws.cell(row=row, column=8, value=status).border = THIN_BORDER

            row += 1

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


def build_quota_alerts_sheet(wb, subs_data):
    """Tab 5: Only quotas at >80% - the action items."""
    ws = wb.create_sheet("⚠ Quota Alerts")

    headers = ["Subscription", "Provider", "Region", "Quota Name", "Limit", "Current Usage", "% Used"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    alerts = []
    for sub_name, data in subs_data:
        quota = data.get("quota-usage")
        if not quota or not quota.get("records"):
            continue
        for rec in quota["records"]:
            pct = rec.get("usagePercent", 0) or 0
            if pct >= 80:
                alerts.append((pct, sub_name, rec))

    # Sort by usage descending
    alerts.sort(key=lambda x: x[0], reverse=True)

    for pct, sub_name, rec in alerts:
        ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
        ws.cell(row=row, column=2, value=rec.get("provider", "")).border = THIN_BORDER
        ws.cell(row=row, column=3, value=rec.get("region", "")).border = THIN_BORDER
        ws.cell(row=row, column=4, value=rec.get("quotaName", "")).border = THIN_BORDER
        ws.cell(row=row, column=5, value=rec.get("limit", 0)).border = THIN_BORDER
        ws.cell(row=row, column=6, value=rec.get("currentUsage", 0)).border = THIN_BORDER

        pct_cell = ws.cell(row=row, column=7, value=round(pct, 1))
        pct_cell.border = THIN_BORDER
        pct_cell.number_format = "0.0"
        pct_cell.fill = CRIT_FILL if pct >= 90 else WARN_FILL

        row += 1

    if row == 2:
        ws.cell(row=2, column=1, value="No quota alerts — all quotas below 80%")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    auto_width(ws)


def build_usage_trends_sheet(wb, subs_data):
    """Tab 6: Usage trends / metrics across subscriptions."""
    ws = wb.create_sheet("Usage Trends")

    headers = [
        "Subscription", "Resource Name", "Resource Type", "Region",
        "Metric", "Unit", "Average", "P95", "Maximum", "Data Points",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    for sub_name, data in subs_data:
        trends = data.get("usage-trends")
        if not trends or not trends.get("records"):
            continue
        for rec in trends["records"]:
            ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=rec.get("resourceName", "")).border = THIN_BORDER
            ws.cell(row=row, column=3, value=rec.get("resourceType", "")).border = THIN_BORDER
            ws.cell(row=row, column=4, value=rec.get("location", rec.get("region", ""))).border = THIN_BORDER
            ws.cell(row=row, column=5, value=rec.get("metricName", "")).border = THIN_BORDER
            ws.cell(row=row, column=6, value=rec.get("unit", "")).border = THIN_BORDER

            for ci, field in enumerate(["average", "p95", "maximum"], 7):
                val = rec.get(field)
                c = ws.cell(row=row, column=ci, value=round(val, 2) if val is not None else "")
                c.border = THIN_BORDER
                c.number_format = "0.00"
                if field == "p95" and rec.get("unit") == "Percent" and val is not None and val >= 80:
                    c.fill = WARN_FILL

            ws.cell(row=row, column=10, value=rec.get("dataPointCount", "")).border = THIN_BORDER
            row += 1

    auto_width(ws)
    ws.auto_filter.ref = ws.dimensions


def build_reservations_sheet(wb, subs_data):
    """Tab 7: Reserved instances across all subscriptions."""
    ws = wb.create_sheet("Reservations")

    headers = [
        "Subscription", "Name", "Type", "SKU", "Region",
        "Quantity", "Term", "Scope", "State",
        "Effective Date", "Expiry Date", "Days Until Expiry",
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "A2"

    row = 2
    any_data = False
    for sub_name, data in subs_data:
        ri = data.get("reserved-instances")
        if not ri or not ri.get("records"):
            continue
        for rec in ri["records"]:
            any_data = True
            ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
            ws.cell(row=row, column=2, value=rec.get("displayName", "")).border = THIN_BORDER
            ws.cell(row=row, column=3, value=rec.get("reservationType", "")).border = THIN_BORDER
            ws.cell(row=row, column=4, value=rec.get("sku", "")).border = THIN_BORDER
            ws.cell(row=row, column=5, value=rec.get("region", "")).border = THIN_BORDER
            ws.cell(row=row, column=6, value=rec.get("quantity", "")).border = THIN_BORDER
            ws.cell(row=row, column=7, value=rec.get("term", "")).border = THIN_BORDER
            ws.cell(row=row, column=8, value=rec.get("scope", "")).border = THIN_BORDER
            ws.cell(row=row, column=9, value=rec.get("provisioningState", "")).border = THIN_BORDER
            ws.cell(row=row, column=10, value=rec.get("effectiveDateTime", "")).border = THIN_BORDER
            ws.cell(row=row, column=11, value=rec.get("expiryDate", "")).border = THIN_BORDER

            days = rec.get("daysUntilExpiry")
            c_days = ws.cell(row=row, column=12, value=days if days is not None else "")
            c_days.border = THIN_BORDER
            if days is not None and days <= 90:
                c_days.fill = CRIT_FILL

            row += 1

    if not any_data:
        ws.cell(row=2, column=1, value="No reservation data available (may require billing permissions)")
        ws.cell(row=2, column=1).font = Font(italic=True, color="666666")

    auto_width(ws)


def build_region_summary_sheet(wb, subs_data):
    """Tab 8: Region usage summary - which subs use which regions."""
    ws = wb.create_sheet("Region Usage")

    # Collect region usage across subs
    region_usage = {}  # {region: {sub: count}}
    all_subs = []
    for sub_name, data in subs_data:
        inv = data.get("service-inventory")
        if not inv or not inv.get("summary"):
            continue
        all_subs.append(sub_name)
        top_regions = inv["summary"].get("topRegions", [])
        for r in top_regions:
            region = r.get("region", "")
            if region not in region_usage:
                region_usage[region] = {}
            region_usage[region][sub_name] = r.get("count", 0)

    if not region_usage:
        ws.cell(row=1, column=1, value="No region data available")
        return

    sorted_regions = sorted(region_usage.keys())

    headers = ["Subscription"] + sorted_regions + ["Total"]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    style_header_row(ws, 1, len(headers))
    ws.freeze_panes = "B2"

    row = 2
    for sub_name in all_subs:
        ws.cell(row=row, column=1, value=sub_name).border = THIN_BORDER
        total = 0
        for col, region in enumerate(sorted_regions, 2):
            val = region_usage.get(region, {}).get(sub_name, 0)
            total += val
            c = ws.cell(row=row, column=col, value=val if val else "")
            c.border = THIN_BORDER
        ws.cell(row=row, column=len(sorted_regions) + 2, value=total).border = THIN_BORDER
        row += 1

    # Totals row
    ws.cell(row=row, column=1, value="TOTAL").font = Font(bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    grand_total = 0
    for col, region in enumerate(sorted_regions, 2):
        region_total = sum(region_usage[region].values())
        grand_total += region_total
        c = ws.cell(row=row, column=col, value=region_total)
        c.font = Font(bold=True)
        c.border = THIN_BORDER
    c = ws.cell(row=row, column=len(sorted_regions) + 2, value=grand_total)
    c.font = Font(bold=True)
    c.border = THIN_BORDER

    auto_width(ws, min_width=6, max_width=25)


def main():
    if len(sys.argv) < 2:
        print("Usage: python3 convert_to_excel.py <report-directory> [output.xlsx]")
        sys.exit(1)

    report_dir = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else os.path.join(report_dir, "capacity-report.xlsx")

    if not os.path.isdir(report_dir):
        print(f"Error: '{report_dir}' is not a directory")
        sys.exit(1)

    # Discover subscriptions
    subs = find_sub_dirs(report_dir)
    if not subs:
        print(f"Error: No subscription data found in '{report_dir}'")
        sys.exit(1)

    print(f"Found {len(subs)} subscription(s)")

    # Load data for each subscription
    json_files = [
        "service-inventory.json",
        "quota-usage.json",
        "usage-trends.json",
        "reserved-instances.json",
    ]

    subs_data = []
    for sub_name, sub_path in subs:
        data = {}
        for jf in json_files:
            key = jf.replace(".json", "")
            data[key] = load_json(os.path.join(sub_path, jf))
        subs_data.append((sub_name, data))
        loaded = sum(1 for v in data.values() if v is not None)
        print(f"  {sub_name}: {loaded}/{len(json_files)} files loaded")

    # Build workbook
    print("\nBuilding Excel workbook...")
    wb = Workbook()

    build_overview_sheet(wb, subs_data)
    print("  ✓ Overview")

    build_inventory_sheet(wb, subs_data)
    print("  ✓ Resource Inventory")

    build_inventory_summary_sheet(wb, subs_data)
    print("  ✓ Resource Summary")

    build_region_summary_sheet(wb, subs_data)
    print("  ✓ Region Usage")

    build_quota_sheet(wb, subs_data)
    print("  ✓ Quota Usage")

    build_quota_alerts_sheet(wb, subs_data)
    print("  ✓ Quota Alerts")

    build_usage_trends_sheet(wb, subs_data)
    print("  ✓ Usage Trends")

    build_reservations_sheet(wb, subs_data)
    print("  ✓ Reservations")

    wb.save(output_file)
    size_kb = os.path.getsize(output_file) / 1024
    print(f"\n✅ Saved: {output_file} ({size_kb:.1f} KB)")


if __name__ == "__main__":
    main()
